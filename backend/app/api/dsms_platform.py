import csv
import re
from datetime import datetime, timezone
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, col, delete, func, or_, select

from app.api.dsms_helpers import (
    generate_temporary_password,
    get_space_or_404,
    page_limit,
    suggest_identifier_key_from_seed,
)
from app.core.config import settings
from app.core.deps import (
    get_current_user,
    get_membership,
    require_super_admin,
    require_tenant_admin,
    require_tenant_member,
)
from app.core.database import get_session
from app.core.security import get_password_hash
from app.core.upload_utils import read_upload_with_limit
from app.models import (
    ProjectSpace,
    QuestionnaireQuestion,
    Tenant,
    TenantCreatorAllowlist,
    TenantMembership,
    User,
)
from app.schemas import (
    PlatformUsersBatchDeactivateIn,
    PlatformUsersPlatformRoleIn,
    MemberBatchIn,
    MemberBatchRemoveIn,
    MemberRoleUpdateIn,
    Page,
    SpaceCreateIn,
    SpaceDeleteIn,
    SpaceOut,
    SpaceUpdateIn,
    TenantCreateIn,
    TenantCreatorUpdate,
    TenantOut,
    TenantPatchIn,
    UserImportItem,
    UserDirectoryOut,
)

router = APIRouter(prefix="/api/v1/dsms", tags=["dsms-platform"])

@router.get(
    "/platform/tenant-creators",
    description="权限：super_admin。获取可创建项目的用户 ID 白名单。",
)
def get_tenant_creators(
    _: User = Depends(require_super_admin), session: Session = Depends(get_session)
):
    ids = session.exec(select(TenantCreatorAllowlist.user_id)).all()
    return {"user_ids": ids}


@router.put(
    "/platform/tenant-creators",
    description="权限：super_admin。全量替换 tenant_creator_allowlist。",
)
def put_tenant_creators(
    payload: TenantCreatorUpdate,
    _: User = Depends(require_super_admin),
    session: Session = Depends(get_session),
):
    session.exec(delete(TenantCreatorAllowlist))
    for user_id in sorted(set(payload.user_ids)):
        if not session.get(User, user_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"用户不存在: {user_id}")
        session.add(TenantCreatorAllowlist(user_id=user_id))
    session.commit()
    return {"user_ids": payload.user_ids, "behavior_key": "dsms-platform-tenant-creators-update"}


@router.get(
    "/platform/users/import-excel/template",
    description="权限：super_admin。下载用户导入模板（xlsx）。",
)
def download_user_import_template(_: User = Depends(require_super_admin)):
    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(["email", "username", "full_name", "department"])
    ws.append(["user1@example.com", "user1", "张三", "数据治理部"])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dsms_users_import_template.xlsx"},
    )


@router.post(
    "/platform/users/import-excel",
    description="权限：super_admin。Excel 批量导入用户；支持 email/邮箱 列名。",
)
def import_users_excel(
    file: UploadFile = File(...),
    _: User = Depends(require_super_admin),
    session: Session = Depends(get_session),
):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="仅支持 .xlsx 文件")

    try:
        raw = read_upload_with_limit(file)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(exc)) from exc
    wb = load_workbook(filename=BytesIO(raw))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Excel 文件为空")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    header_map = {name: idx for idx, name in enumerate(header)}
    email_idx = header_map.get("email", header_map.get("邮箱"))
    username_idx = header_map.get("username", header_map.get("用户名"))
    full_name_idx = header_map.get("full_name", header_map.get("姓名"))
    department_idx = header_map.get("department", header_map.get("部门"))
    if email_idx is None:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="模板缺少 email/邮箱 列")

    created_users: list[dict] = []
    skipped_items = []
    total_rows = max(len(rows) - 1, 0)

    for ridx, row in enumerate(rows[1:], start=2):
        email = str(row[email_idx]).strip() if len(row) > email_idx and row[email_idx] else ""
        username = (
            str(row[username_idx]).strip()
            if username_idx is not None and len(row) > username_idx and row[username_idx]
            else ""
        )
        full_name = (
            str(row[full_name_idx]).strip()
            if full_name_idx is not None and len(row) > full_name_idx and row[full_name_idx]
            else None
        )
        department = (
            str(row[department_idx]).strip()
            if department_idx is not None and len(row) > department_idx and row[department_idx]
            else None
        )

        if not email:
            skipped_items.append({"row": ridx, "email": "", "reason": "邮箱为空"})
            continue
        if session.exec(select(User).where(User.email == email)).first():
            skipped_items.append({"row": ridx, "email": email, "reason": "邮箱已存在"})
            continue
        if not username:
            username = email.split("@")[0]
        base_username = username
        suffix = 1
        while session.exec(select(User).where(User.username == username)).first():
            username = f"{base_username}{suffix}"
            suffix += 1

        temp_password = generate_temporary_password()
        user = User(
            username=username,
            email=email,
            hashed_password=get_password_hash(temp_password),
            full_name=full_name,
            department=department,
            is_active=True,
            is_approved=True,
        )
        session.add(user)
        session.flush()
        created_users.append(
            UserImportItem(
                id=user.id,
                username=user.username,
                email=user.email,
                full_name=user.full_name,
                department=user.department,
                is_active=user.is_active,
                created_at=user.created_at,
                temporary_password=temp_password,
            ).model_dump()
        )

    session.commit()
    return {
        "total_rows": total_rows,
        "created_count": len(created_users),
        "skipped_count": len(skipped_items),
        "created_users": created_users,
        "skipped_items": skipped_items,
        "behavior_key": "dsms-platform-users-import-excel",
    }


@router.post(
    "/platform/users/batch-deactivate",
    description="权限：super_admin。批量停用用户账号（不可停用超管与当前登录账号）。",
)
def batch_deactivate_users(
    payload: PlatformUsersBatchDeactivateIn,
    current_user: User = Depends(require_super_admin),
    session: Session = Depends(get_session),
):
    deactivated: list[int] = []
    skipped_items: list[dict] = []
    for user_id in payload.user_ids:
        user = session.get(User, user_id)
        if not user:
            skipped_items.append({"user_id": user_id, "reason": "用户不存在"})
            continue
        if user.id == current_user.id:
            skipped_items.append({"user_id": user_id, "reason": "不能停用当前登录账号"})
            continue
        if user.is_superuser:
            skipped_items.append({"user_id": user_id, "reason": "不能停用超级管理员"})
            continue
        if not user.is_active:
            skipped_items.append({"user_id": user_id, "reason": "账号已停用"})
            continue
        user.is_active = False
        user.refresh_token_version = (user.refresh_token_version or 0) + 1
        session.add(user)
        deactivated.append(user_id)
    session.commit()
    return {
        "deactivated_user_ids": deactivated,
        "skipped_items": skipped_items,
        "behavior_key": "dsms-platform-users-batch-deactivate",
    }


@router.put(
    "/platform/users/batch-platform-role",
    description="权限：super_admin。批量设置门户 platform_role（不可修改超级管理员）。",
)
def batch_set_platform_role(
    payload: PlatformUsersPlatformRoleIn,
    _: User = Depends(require_super_admin),
    session: Session = Depends(get_session),
):
    allowed = {"system_admin", "security_fo", "function_fo"}
    role = (payload.platform_role or "").strip()
    if role not in allowed:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="无效的平台角色")

    updated: list[int] = []
    skipped_items: list[dict] = []
    for user_id in payload.user_ids:
        user = session.get(User, user_id)
        if not user:
            skipped_items.append({"user_id": user_id, "reason": "用户不存在"})
            continue
        if user.is_superuser:
            skipped_items.append({"user_id": user_id, "reason": "超级管理员不可修改平台角色"})
            continue
        user.platform_role = role
        session.add(user)
        updated.append(user_id)
    session.commit()
    return {
        "updated_user_ids": updated,
        "skipped_items": skipped_items,
        "platform_role": role,
        "behavior_key": "dsms-platform-users-batch-platform-role",
    }


@router.get(
    "/tenants",
    response_model=Page,
    description="权限：super_admin / tenant_admin / tenant_member（已登录）。返回当前用户可见项目。",
)
def list_tenants(
    skip: int = 0,
    limit: int = Query(default=20),
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    if current_user.is_superuser:
        stmt = select(Tenant)
    else:
        tenant_ids = select(TenantMembership.tenant_id).where(TenantMembership.user_id == current_user.id)
        stmt = select(Tenant).where(col(Tenant.id).in_(tenant_ids))
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    items = session.exec(stmt.order_by(Tenant.id).offset(skip).limit(limit)).all()
    return {"total": total, "items": [TenantOut.model_validate(i, from_attributes=True).model_dump() for i in items]}


@router.get(
    "/users",
    response_model=Page,
    description="权限：super_admin 或 tenant_admin。返回全局用户目录（白名单字段）。",
)
def list_users(
    skip: int = 0,
    limit: int = Query(default=20),
    q: str | None = None,
    is_active: bool | None = None,
    membership_preview_tenant_id: int | None = None,
    only_unassigned_to_tenant: int | None = None,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    is_admin_any_tenant = session.exec(
        select(TenantMembership).where(
            TenantMembership.user_id == current_user.id, TenantMembership.role == "tenant_admin"
        )
    ).first()
    if not current_user.is_superuser and not is_admin_any_tenant:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="仅超管或项目管理员可访问用户目录")

    stmt = select(User)
    if (
        not current_user.is_superuser
        and settings.tenant_admin_user_directory_scope.strip().lower() != "all"
    ):
        admin_tenant_ids = select(TenantMembership.tenant_id).where(
            TenantMembership.user_id == current_user.id,
            TenantMembership.role == "tenant_admin",
        )
        scoped_user_ids = select(TenantMembership.user_id).where(
            col(TenantMembership.tenant_id).in_(admin_tenant_ids)
        )
        stmt = stmt.where(col(User.id).in_(scoped_user_ids))
    if q:
        like_val = f"%{q}%"
        stmt = stmt.where(or_(User.username.like(like_val), User.email.like(like_val), User.full_name.like(like_val)))
    if is_active is not None:
        stmt = stmt.where(User.is_active == is_active)
    if only_unassigned_to_tenant:
        assigned_subquery = select(TenantMembership.user_id).where(TenantMembership.tenant_id == only_unassigned_to_tenant)
        stmt = stmt.where(col(User.id).not_in(assigned_subquery))

    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    users = session.exec(stmt.order_by(User.id).offset(skip).limit(limit)).all()
    items = []
    for user in users:
        item = UserDirectoryOut(
            id=user.id,
            username=user.username,
            email=user.email,
            full_name=user.full_name,
            department=user.department,
            is_active=user.is_active,
            is_superuser=user.is_superuser,
            platform_role=user.platform_role or "security_fo",
            created_at=user.created_at,
        )
        if membership_preview_tenant_id:
            m = get_membership(session, membership_preview_tenant_id, user.id)
            item.in_tenant = m is not None
            item.tenant_role = m.role if m else None
        items.append(item.model_dump())
    return {"total": total, "items": items}


@router.post(
    "/tenants",
    description="权限：super_admin 或 tenant_creator（allowlist）。创建项目并把创建者写为 tenant_admin。",
)
def create_tenant(
    payload: TenantCreateIn,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    if not current_user.is_superuser:
        allowed = session.get(TenantCreatorAllowlist, current_user.id)
        if not allowed:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="您没有创建项目的权限")
    tenant = Tenant(name=payload.name, slug=payload.slug)
    session.add(tenant)
    session.flush()
    session.add(TenantMembership(tenant_id=tenant.id, user_id=current_user.id, role="tenant_admin"))
    session.commit()
    session.refresh(tenant)
    return tenant


@router.get(
    "/tenants/{tenant_id}",
    response_model=TenantOut,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。查看项目详情。",
)
def get_tenant(
    tenant_id: int,
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    tenant = session.get(Tenant, tenant_id)
    if not tenant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")
    return TenantOut.model_validate(tenant, from_attributes=True)


@router.patch(
    "/tenants/{tenant_id}",
    response_model=TenantOut,
    description="权限：tenant_admin（同项目）或 super_admin。更新项目基础信息。",
)
def patch_tenant(
    tenant_id: int,
    payload: TenantPatchIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    tenant = session.get(Tenant, tenant_id)
    if not tenant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")
    data = payload.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(tenant, key, value)
    session.add(tenant)
    session.commit()
    session.refresh(tenant)
    return TenantOut.model_validate(tenant, from_attributes=True)


@router.delete(
    "/tenants/{tenant_id}",
    description="权限：super_admin。删除项目及其成员与全部空间数据（不可恢复）；默认项目（slug=default）不可删。",
)
def delete_tenant(
    tenant_id: int,
    _: User = Depends(require_super_admin),
    session: Session = Depends(get_session),
):
    from app.services.tenant_delete_service import delete_tenant_cascade

    result = delete_tenant_cascade(session, tenant_id)
    if not result.get("ok"):
        reason = result.get("reason")
        if reason == "not_found":
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=result.get("message"))
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=result.get("message"))
    session.commit()
    return {"deleted_tenant_id": tenant_id, "behavior_key": "dsms-tenant-delete"}


@router.post(
    "/tenants/{tenant_id}/seeds/import",
    description="权限：tenant_admin（同项目）或 super_admin。导入最小治理种子（空间+问卷题目）。",
)
def import_tenant_seeds(
    tenant_id: int,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    tenant = session.get(Tenant, tenant_id)
    if not tenant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")

    created = {"spaces": 0, "questions": 0}
    skipped = {"spaces": 0, "questions": 0}

    seed_space_key = "default-space"
    space = session.exec(
        select(ProjectSpace).where(ProjectSpace.tenant_id == tenant_id, ProjectSpace.space_key == seed_space_key)
    ).first()
    if not space:
        space = ProjectSpace(tenant_id=tenant_id, space_key=seed_space_key, name="默认项目空间")
        session.add(space)
        session.flush()
        created["spaces"] += 1
    else:
        skipped["spaces"] += 1

    default_questions = [
        {"key": "business_scope", "title": "该业务功能是否涉及个人信息处理？", "sort_order": 1},
        {"key": "cross_border", "title": "该业务功能是否涉及跨境数据传输？", "sort_order": 2},
    ]
    for q in default_questions:
        existing = session.exec(
            select(QuestionnaireQuestion).where(
                QuestionnaireQuestion.tenant_id == tenant_id,
                QuestionnaireQuestion.project_space_id == space.id,
                QuestionnaireQuestion.key == q["key"],
            )
        ).first()
        if existing:
            skipped["questions"] += 1
            continue
        session.add(
            QuestionnaireQuestion(
                tenant_id=tenant_id,
                project_space_id=space.id,
                key=q["key"],
                title=q["title"],
                question_type="text",
                is_required=False,
                sort_order=q["sort_order"],
            )
        )
        created["questions"] += 1

    session.commit()
    return {
        "tenant_id": tenant_id,
        "space_id": space.id,
        "created": created,
        "skipped": skipped,
        "behavior_key": "dsms-seeds-import",
    }


@router.get(
    "/tenants/{tenant_id}/members",
    description="权限：tenant_admin（同项目）或 super_admin。查看项目成员。",
)
def list_members(
    tenant_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    stmt = (
        select(TenantMembership, User)
        .join(User, User.id == TenantMembership.user_id)
        .where(TenantMembership.tenant_id == tenant_id)
    )
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    rows = session.exec(stmt.order_by(TenantMembership.id).offset(skip).limit(limit)).all()
    items = [
        {
            "user_id": user.id,
            "username": user.username,
            "email": user.email,
            "role": membership.role,
            "created_at": membership.created_at,
        }
        for membership, user in rows
    ]
    return {"total": total, "items": items}


@router.post(
    "/tenants/{tenant_id}/members/batch",
    description="权限：tenant_admin（同项目）或 super_admin。批量加入成员，默认 tenant_member。",
)
def batch_add_members(
    tenant_id: int,
    payload: MemberBatchIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    if payload.role != "tenant_member":
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="批量加入仅允许 role=tenant_member")
    added, skipped = [], []
    for user_id in payload.user_ids:
        if not session.get(User, user_id):
            skipped.append({"user_id": user_id, "reason": "用户不存在"})
            continue
        exists = get_membership(session, tenant_id, user_id)
        if exists:
            skipped.append({"user_id": user_id, "reason": "成员已存在"})
            continue
        membership = TenantMembership(tenant_id=tenant_id, user_id=user_id, role="tenant_member")
        session.add(membership)
        added.append(user_id)
    session.commit()
    return {"added_user_ids": added, "skipped_items": skipped, "behavior_key": "dsms-tenant-members-batch-add"}


@router.post(
    "/tenants/{tenant_id}/members/batch-remove",
    description="权限：tenant_admin（同项目）或 super_admin。批量移除成员，且保留至少 1 名 tenant_admin。",
)
def batch_remove_members(
    tenant_id: int,
    payload: MemberBatchRemoveIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    removed, skipped = [], []
    for user_id in payload.user_ids:
        m = get_membership(session, tenant_id, user_id)
        if not m:
            skipped.append({"user_id": user_id, "reason": "成员不存在"})
            continue
        if m.role == "tenant_admin":
            admin_count = session.exec(
                select(func.count()).where(
                    TenantMembership.tenant_id == tenant_id, TenantMembership.role == "tenant_admin"
                )
            ).one()
            if admin_count <= 1:
                skipped.append({"user_id": user_id, "reason": "至少保留一名项目管理员"})
                continue
        session.delete(m)
        removed.append(user_id)
    session.commit()
    return {"removed_user_ids": removed, "skipped_items": skipped, "behavior_key": "dsms-tenant-members-batch-remove"}


@router.put(
    "/tenants/{tenant_id}/members/{user_id}/role",
    description="权限：super_admin 或同项目 tenant_admin。角色仅 tenant_admin/tenant_member。",
)
def update_member_role(
    tenant_id: int,
    user_id: int,
    payload: MemberRoleUpdateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    if payload.role not in {"tenant_admin", "tenant_member"}:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="角色仅支持 tenant_admin/tenant_member")
    membership = get_membership(session, tenant_id, user_id)
    if not membership:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="成员不存在")
    if membership.role == "tenant_admin" and payload.role == "tenant_member":
        admin_count = session.exec(
            select(func.count()).where(TenantMembership.tenant_id == tenant_id, TenantMembership.role == "tenant_admin")
        ).one()
        if admin_count <= 1:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="至少保留一名项目管理员")
    membership.role = payload.role
    session.add(membership)
    session.commit()
    return {"tenant_id": tenant_id, "user_id": user_id, "role": payload.role, "behavior_key": "dsms-tenant-member-role-promote"}


@router.get(
    "/tenants/{tenant_id}/spaces",
    response_model=Page,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。项目空间分页列表。",
)
def list_spaces(
    tenant_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    stmt = select(ProjectSpace).where(ProjectSpace.tenant_id == tenant_id)
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    items = session.exec(stmt.order_by(ProjectSpace.id).offset(skip).limit(limit)).all()
    return {"total": total, "items": [SpaceOut.model_validate(i, from_attributes=True).model_dump() for i in items]}


@router.post(
    "/tenants/{tenant_id}/spaces",
    response_model=SpaceOut,
    description="权限：tenant_admin（同项目）或 super_admin。创建项目空间。",
)
def create_space(
    tenant_id: int,
    payload: SpaceCreateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    exists = session.exec(
        select(ProjectSpace).where(ProjectSpace.tenant_id == tenant_id, ProjectSpace.space_key == payload.space_key)
    ).first()
    if exists:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="space_key 已存在")
    item = ProjectSpace(tenant_id=tenant_id, space_key=payload.space_key, name=payload.name)
    session.add(item)
    session.commit()
    session.refresh(item)
    return SpaceOut.model_validate(item, from_attributes=True)


@router.put(
    "/tenants/{tenant_id}/spaces",
    response_model=SpaceOut,
    description="权限：tenant_admin（同项目）或 super_admin。更新项目空间（统一使用 PUT /spaces）。",
)
def update_space(
    tenant_id: int,
    payload: SpaceUpdateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    item = session.get(ProjectSpace, payload.id)
    if not item or item.tenant_id != tenant_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="空间不存在")
    item.name = payload.name
    item.updated_at = datetime.now(timezone.utc)
    session.add(item)
    session.commit()
    session.refresh(item)
    return SpaceOut.model_validate(item, from_attributes=True)


@router.post(
    "/tenants/{tenant_id}/spaces/delete",
    description="权限：tenant_admin（同项目）或 super_admin。批量删除项目空间。",
)
def delete_spaces(
    tenant_id: int,
    payload: SpaceDeleteIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    for sid in payload.ids:
        space = session.get(ProjectSpace, sid)
        if space and space.tenant_id == tenant_id:
            session.delete(space)
    session.commit()
    return {"deleted_ids": payload.ids, "behavior_key": "project-spaces/delete"}


