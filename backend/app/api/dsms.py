import csv
import re
from datetime import datetime, timezone
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, col, delete, func, or_, select

from app.core.deps import (
    get_current_user,
    get_membership,
    require_super_admin,
    require_tenant_admin,
    require_tenant_member,
)
from app.core.database import get_session
from app.core.security import get_password_hash
from app.models import (
    BusinessFunctionOption,
    BusinessFunctionOptionRequest,
    FieldCatalogEntry,
    FieldRequest,
    FieldUsageReport,
    FieldUsageReportItem,
    LifecycleFieldConfig,
    ProjectSpace,
    QuestionnaireQuestion,
    RelevanceAssessmentAnswer,
    RelevanceAssessmentSubmission,
    RelevanceRule,
    SensitivityLevel,
    TaxonomyLevel,
    Tenant,
    TenantCreatorAllowlist,
    TenantMembership,
    TaxonomyNode,
    User,
)
from app.schemas import (
    PlatformUsersBatchDeactivateIn,
    PlatformUsersPlatformRoleIn,
    BusinessFunctionOptionOut,
    BusinessFunctionOptionRequestCreateIn,
    BusinessFunctionOptionRequestOut,
    BusinessFunctionOptionRequestReviewIn,
    FieldCatalogCreateIn,
    FieldCatalogOut,
    FieldCatalogUpdateIn,
    FieldUsageReportCreateIn,
    FieldUsageReportItemOut,
    FieldUsageReportListItemOut,
    FieldUsageReportOut,
    FieldUsageReportReviewIn,
    FieldRequestCreateIn,
    FieldRequestOut,
    FieldRequestReviewIn,
    IdentifierSuggestIn,
    IdentifierSuggestOut,
    LifecycleFieldConfigCreateIn,
    LifecycleFieldConfigOut,
    LifecycleFieldConfigUpdateIn,
    MemberBatchIn,
    MemberBatchRemoveIn,
    MemberRoleUpdateIn,
    Page,
    QuestionCreateIn,
    QuestionDeleteIn,
    QuestionOut,
    QuestionUpdateIn,
    RelevanceAssessmentCreateIn,
    RelevanceAssessmentOut,
    RelevanceAssessmentAnswerIn,
    RelevanceRuleIn,
    RelevanceRuleOut,
    SensitivityLevelOut,
    TaxonomyLevelCreateIn,
    TaxonomyLevelOut,
    TaxonomyLevelUpdateIn,
    SpaceCreateIn,
    SpaceDeleteIn,
    SpaceOut,
    SpaceUpdateIn,
    TenantCreateIn,
    TenantCreatorUpdate,
    TenantOut,
    TenantPatchIn,
    TaxonomyNodeCreateIn,
    TaxonomyNodeOut,
    TaxonomyNodeUpdateIn,
    UserImportItem,
    UserDirectoryOut,
)

router = APIRouter(prefix="/api/v1/dsms", tags=["dsms"])


def page_limit(limit: int) -> int:
    if limit > 500:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="limit 最大值为 500")
    return limit


def suggest_identifier_key_from_seed(seed: str) -> str:
    s = seed.strip().lower()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s)
    if not s:
        h = abs(hash(seed)) % (10**8)
        s = f"field_{h}"
    return s[:80]


@router.get(
    "/platform/tenant-creators",
    description="权限：super_admin。获取可创建项目的用户 ID 白名单。",
)
def get_tenant_creators(
    _: User = Depends(require_super_admin), session: Session = Depends(get_session)
):
    ids = session.exec(select(TenantCreatorAllowlist.user_id)).all()
    return {"user_ids": ids}


@router.put(
    "/platform/tenant-creators",
    description="权限：super_admin。全量替换 tenant_creator_allowlist。",
)
def put_tenant_creators(
    payload: TenantCreatorUpdate,
    _: User = Depends(require_super_admin),
    session: Session = Depends(get_session),
):
    session.exec(delete(TenantCreatorAllowlist))
    for user_id in sorted(set(payload.user_ids)):
        if not session.get(User, user_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"用户不存在: {user_id}")
        session.add(TenantCreatorAllowlist(user_id=user_id))
    session.commit()
    return {"user_ids": payload.user_ids, "behavior_key": "dsms-platform-tenant-creators-update"}


@router.get(
    "/platform/users/import-excel/template",
    description="权限：super_admin。下载用户导入模板（xlsx）。",
)
def download_user_import_template(_: User = Depends(require_super_admin)):
    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(["email", "username", "full_name", "department"])
    ws.append(["user1@example.com", "user1", "张三", "数据治理部"])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dsms_users_import_template.xlsx"},
    )


@router.post(
    "/platform/users/import-excel",
    description="权限：super_admin。Excel 批量导入用户；支持 email/邮箱 列名。",
)
def import_users_excel(
    file: UploadFile = File(...),
    _: User = Depends(require_super_admin),
    session: Session = Depends(get_session),
):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="仅支持 .xlsx 文件")

    wb = load_workbook(filename=BytesIO(file.file.read()))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Excel 文件为空")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    header_map = {name: idx for idx, name in enumerate(header)}
    email_idx = header_map.get("email", header_map.get("邮箱"))
    username_idx = header_map.get("username", header_map.get("用户名"))
    full_name_idx = header_map.get("full_name", header_map.get("姓名"))
    department_idx = header_map.get("department", header_map.get("部门"))
    if email_idx is None:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="模板缺少 email/邮箱 列")

    created_users: list[dict] = []
    skipped_items = []
    total_rows = max(len(rows) - 1, 0)

    for ridx, row in enumerate(rows[1:], start=2):
        email = str(row[email_idx]).strip() if len(row) > email_idx and row[email_idx] else ""
        username = (
            str(row[username_idx]).strip()
            if username_idx is not None and len(row) > username_idx and row[username_idx]
            else ""
        )
        full_name = (
            str(row[full_name_idx]).strip()
            if full_name_idx is not None and len(row) > full_name_idx and row[full_name_idx]
            else None
        )
        department = (
            str(row[department_idx]).strip()
            if department_idx is not None and len(row) > department_idx and row[department_idx]
            else None
        )

        if not email:
            skipped_items.append({"row": ridx, "email": "", "reason": "邮箱为空"})
            continue
        if session.exec(select(User).where(User.email == email)).first():
            skipped_items.append({"row": ridx, "email": email, "reason": "邮箱已存在"})
            continue
        if not username:
            username = email.split("@")[0]
        base_username = username
        suffix = 1
        while session.exec(select(User).where(User.username == username)).first():
            username = f"{base_username}{suffix}"
            suffix += 1

        user = User(
            username=username,
            email=email,
            hashed_password=get_password_hash("Admin123456"),
            full_name=full_name,
            department=department,
            is_active=True,
            is_approved=True,
        )
        session.add(user)
        session.flush()
        created_users.append(
            UserImportItem(
                id=user.id,
                username=user.username,
                email=user.email,
                full_name=user.full_name,
                department=user.department,
                is_active=user.is_active,
                created_at=user.created_at,
            ).model_dump()
        )

    session.commit()
    return {
        "total_rows": total_rows,
        "created_count": len(created_users),
        "skipped_count": len(skipped_items),
        "created_users": created_users,
        "skipped_items": skipped_items,
        "behavior_key": "dsms-platform-users-import-excel",
    }


@router.post(
    "/platform/users/batch-deactivate",
    description="权限：super_admin。批量停用用户账号（不可停用超管与当前登录账号）。",
)
def batch_deactivate_users(
    payload: PlatformUsersBatchDeactivateIn,
    current_user: User = Depends(require_super_admin),
    session: Session = Depends(get_session),
):
    deactivated: list[int] = []
    skipped_items: list[dict] = []
    for user_id in payload.user_ids:
        user = session.get(User, user_id)
        if not user:
            skipped_items.append({"user_id": user_id, "reason": "用户不存在"})
            continue
        if user.id == current_user.id:
            skipped_items.append({"user_id": user_id, "reason": "不能停用当前登录账号"})
            continue
        if user.is_superuser:
            skipped_items.append({"user_id": user_id, "reason": "不能停用超级管理员"})
            continue
        if not user.is_active:
            skipped_items.append({"user_id": user_id, "reason": "账号已停用"})
            continue
        user.is_active = False
        session.add(user)
        deactivated.append(user_id)
    session.commit()
    return {
        "deactivated_user_ids": deactivated,
        "skipped_items": skipped_items,
        "behavior_key": "dsms-platform-users-batch-deactivate",
    }


@router.put(
    "/platform/users/batch-platform-role",
    description="权限：super_admin。批量设置门户 platform_role（不可修改超级管理员）。",
)
def batch_set_platform_role(
    payload: PlatformUsersPlatformRoleIn,
    _: User = Depends(require_super_admin),
    session: Session = Depends(get_session),
):
    allowed = {"system_admin", "security_fo", "function_fo"}
    role = (payload.platform_role or "").strip()
    if role not in allowed:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="无效的平台角色")

    updated: list[int] = []
    skipped_items: list[dict] = []
    for user_id in payload.user_ids:
        user = session.get(User, user_id)
        if not user:
            skipped_items.append({"user_id": user_id, "reason": "用户不存在"})
            continue
        if user.is_superuser:
            skipped_items.append({"user_id": user_id, "reason": "超级管理员不可修改平台角色"})
            continue
        user.platform_role = role
        session.add(user)
        updated.append(user_id)
    session.commit()
    return {
        "updated_user_ids": updated,
        "skipped_items": skipped_items,
        "platform_role": role,
        "behavior_key": "dsms-platform-users-batch-platform-role",
    }


@router.get(
    "/tenants",
    response_model=Page,
    description="权限：super_admin / tenant_admin / tenant_member（已登录）。返回当前用户可见项目。",
)
def list_tenants(
    skip: int = 0,
    limit: int = Query(default=20),
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    if current_user.is_superuser:
        stmt = select(Tenant)
    else:
        tenant_ids = select(TenantMembership.tenant_id).where(TenantMembership.user_id == current_user.id)
        stmt = select(Tenant).where(col(Tenant.id).in_(tenant_ids))
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    items = session.exec(stmt.order_by(Tenant.id).offset(skip).limit(limit)).all()
    return {"total": total, "items": [TenantOut.model_validate(i, from_attributes=True).model_dump() for i in items]}


@router.get(
    "/users",
    response_model=Page,
    description="权限：super_admin 或 tenant_admin。返回全局用户目录（白名单字段）。",
)
def list_users(
    skip: int = 0,
    limit: int = Query(default=20),
    q: str | None = None,
    is_active: bool | None = None,
    membership_preview_tenant_id: int | None = None,
    only_unassigned_to_tenant: int | None = None,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    is_admin_any_tenant = session.exec(
        select(TenantMembership).where(
            TenantMembership.user_id == current_user.id, TenantMembership.role == "tenant_admin"
        )
    ).first()
    if not current_user.is_superuser and not is_admin_any_tenant:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="仅超管或项目管理员可访问用户目录")

    stmt = select(User)
    if q:
        like_val = f"%{q}%"
        stmt = stmt.where(or_(User.username.like(like_val), User.email.like(like_val), User.full_name.like(like_val)))
    if is_active is not None:
        stmt = stmt.where(User.is_active == is_active)
    if only_unassigned_to_tenant:
        assigned_subquery = select(TenantMembership.user_id).where(TenantMembership.tenant_id == only_unassigned_to_tenant)
        stmt = stmt.where(col(User.id).not_in(assigned_subquery))

    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    users = session.exec(stmt.order_by(User.id).offset(skip).limit(limit)).all()
    items = []
    for user in users:
        item = UserDirectoryOut(
            id=user.id,
            username=user.username,
            email=user.email,
            full_name=user.full_name,
            department=user.department,
            is_active=user.is_active,
            is_superuser=user.is_superuser,
            platform_role=user.platform_role or "security_fo",
            created_at=user.created_at,
        )
        if membership_preview_tenant_id:
            m = get_membership(session, membership_preview_tenant_id, user.id)
            item.in_tenant = m is not None
            item.tenant_role = m.role if m else None
        items.append(item.model_dump())
    return {"total": total, "items": items}


@router.post(
    "/tenants",
    description="权限：super_admin 或 tenant_creator（allowlist）。创建项目并把创建者写为 tenant_admin。",
)
def create_tenant(
    payload: TenantCreateIn,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    if not current_user.is_superuser:
        allowed = session.get(TenantCreatorAllowlist, current_user.id)
        if not allowed:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="您没有创建项目的权限")
    tenant = Tenant(name=payload.name, slug=payload.slug)
    session.add(tenant)
    session.flush()
    session.add(TenantMembership(tenant_id=tenant.id, user_id=current_user.id, role="tenant_admin"))
    session.commit()
    session.refresh(tenant)
    return tenant


@router.get(
    "/tenants/{tenant_id}",
    response_model=TenantOut,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。查看项目详情。",
)
def get_tenant(
    tenant_id: int,
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    tenant = session.get(Tenant, tenant_id)
    if not tenant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")
    return TenantOut.model_validate(tenant, from_attributes=True)


@router.patch(
    "/tenants/{tenant_id}",
    response_model=TenantOut,
    description="权限：tenant_admin（同项目）或 super_admin。更新项目基础信息。",
)
def patch_tenant(
    tenant_id: int,
    payload: TenantPatchIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    tenant = session.get(Tenant, tenant_id)
    if not tenant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")
    data = payload.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(tenant, key, value)
    session.add(tenant)
    session.commit()
    session.refresh(tenant)
    return TenantOut.model_validate(tenant, from_attributes=True)


@router.delete(
    "/tenants/{tenant_id}",
    description="权限：super_admin。删除项目及其成员与全部空间数据（不可恢复）；默认项目（slug=default）不可删。",
)
def delete_tenant(
    tenant_id: int,
    _: User = Depends(require_super_admin),
    session: Session = Depends(get_session),
):
    from app.services.tenant_delete_service import delete_tenant_cascade

    result = delete_tenant_cascade(session, tenant_id)
    if not result.get("ok"):
        reason = result.get("reason")
        if reason == "not_found":
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=result.get("message"))
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=result.get("message"))
    session.commit()
    return {"deleted_tenant_id": tenant_id, "behavior_key": "dsms-tenant-delete"}


@router.post(
    "/tenants/{tenant_id}/seeds/import",
    description="权限：tenant_admin（同项目）或 super_admin。导入最小治理种子（空间+问卷题目）。",
)
def import_tenant_seeds(
    tenant_id: int,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    tenant = session.get(Tenant, tenant_id)
    if not tenant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")

    created = {"spaces": 0, "questions": 0}
    skipped = {"spaces": 0, "questions": 0}

    seed_space_key = "default-space"
    space = session.exec(
        select(ProjectSpace).where(ProjectSpace.tenant_id == tenant_id, ProjectSpace.space_key == seed_space_key)
    ).first()
    if not space:
        space = ProjectSpace(tenant_id=tenant_id, space_key=seed_space_key, name="默认项目空间")
        session.add(space)
        session.flush()
        created["spaces"] += 1
    else:
        skipped["spaces"] += 1

    default_questions = [
        {"key": "business_scope", "title": "该业务功能是否涉及个人信息处理？", "sort_order": 1},
        {"key": "cross_border", "title": "该业务功能是否涉及跨境数据传输？", "sort_order": 2},
    ]
    for q in default_questions:
        existing = session.exec(
            select(QuestionnaireQuestion).where(
                QuestionnaireQuestion.tenant_id == tenant_id,
                QuestionnaireQuestion.project_space_id == space.id,
                QuestionnaireQuestion.key == q["key"],
            )
        ).first()
        if existing:
            skipped["questions"] += 1
            continue
        session.add(
            QuestionnaireQuestion(
                tenant_id=tenant_id,
                project_space_id=space.id,
                key=q["key"],
                title=q["title"],
                question_type="text",
                is_required=False,
                sort_order=q["sort_order"],
            )
        )
        created["questions"] += 1

    session.commit()
    return {
        "tenant_id": tenant_id,
        "space_id": space.id,
        "created": created,
        "skipped": skipped,
        "behavior_key": "dsms-seeds-import",
    }


@router.get(
    "/tenants/{tenant_id}/members",
    description="权限：tenant_admin（同项目）或 super_admin。查看项目成员。",
)
def list_members(
    tenant_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    stmt = (
        select(TenantMembership, User)
        .join(User, User.id == TenantMembership.user_id)
        .where(TenantMembership.tenant_id == tenant_id)
    )
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    rows = session.exec(stmt.order_by(TenantMembership.id).offset(skip).limit(limit)).all()
    items = [
        {
            "user_id": user.id,
            "username": user.username,
            "email": user.email,
            "role": membership.role,
            "created_at": membership.created_at,
        }
        for membership, user in rows
    ]
    return {"total": total, "items": items}


@router.post(
    "/tenants/{tenant_id}/members/batch",
    description="权限：tenant_admin（同项目）或 super_admin。批量加入成员，默认 tenant_member。",
)
def batch_add_members(
    tenant_id: int,
    payload: MemberBatchIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    if payload.role != "tenant_member":
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="批量加入仅允许 role=tenant_member")
    added, skipped = [], []
    for user_id in payload.user_ids:
        if not session.get(User, user_id):
            skipped.append({"user_id": user_id, "reason": "用户不存在"})
            continue
        exists = get_membership(session, tenant_id, user_id)
        if exists:
            skipped.append({"user_id": user_id, "reason": "成员已存在"})
            continue
        membership = TenantMembership(tenant_id=tenant_id, user_id=user_id, role="tenant_member")
        session.add(membership)
        added.append(user_id)
    session.commit()
    return {"added_user_ids": added, "skipped_items": skipped, "behavior_key": "dsms-tenant-members-batch-add"}


@router.post(
    "/tenants/{tenant_id}/members/batch-remove",
    description="权限：tenant_admin（同项目）或 super_admin。批量移除成员，且保留至少 1 名 tenant_admin。",
)
def batch_remove_members(
    tenant_id: int,
    payload: MemberBatchRemoveIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    removed, skipped = [], []
    for user_id in payload.user_ids:
        m = get_membership(session, tenant_id, user_id)
        if not m:
            skipped.append({"user_id": user_id, "reason": "成员不存在"})
            continue
        if m.role == "tenant_admin":
            admin_count = session.exec(
                select(func.count()).where(
                    TenantMembership.tenant_id == tenant_id, TenantMembership.role == "tenant_admin"
                )
            ).one()
            if admin_count <= 1:
                skipped.append({"user_id": user_id, "reason": "至少保留一名项目管理员"})
                continue
        session.delete(m)
        removed.append(user_id)
    session.commit()
    return {"removed_user_ids": removed, "skipped_items": skipped, "behavior_key": "dsms-tenant-members-batch-remove"}


@router.put(
    "/tenants/{tenant_id}/members/{user_id}/role",
    description="权限：super_admin 或同项目 tenant_admin。角色仅 tenant_admin/tenant_member。",
)
def update_member_role(
    tenant_id: int,
    user_id: int,
    payload: MemberRoleUpdateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    if payload.role not in {"tenant_admin", "tenant_member"}:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="角色仅支持 tenant_admin/tenant_member")
    membership = get_membership(session, tenant_id, user_id)
    if not membership:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="成员不存在")
    if membership.role == "tenant_admin" and payload.role == "tenant_member":
        admin_count = session.exec(
            select(func.count()).where(TenantMembership.tenant_id == tenant_id, TenantMembership.role == "tenant_admin")
        ).one()
        if admin_count <= 1:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="至少保留一名项目管理员")
    membership.role = payload.role
    session.add(membership)
    session.commit()
    return {"tenant_id": tenant_id, "user_id": user_id, "role": payload.role, "behavior_key": "dsms-tenant-member-role-promote"}


@router.get(
    "/tenants/{tenant_id}/spaces",
    response_model=Page,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。项目空间分页列表。",
)
def list_spaces(
    tenant_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    stmt = select(ProjectSpace).where(ProjectSpace.tenant_id == tenant_id)
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    items = session.exec(stmt.order_by(ProjectSpace.id).offset(skip).limit(limit)).all()
    return {"total": total, "items": [SpaceOut.model_validate(i, from_attributes=True).model_dump() for i in items]}


@router.post(
    "/tenants/{tenant_id}/spaces",
    response_model=SpaceOut,
    description="权限：tenant_admin（同项目）或 super_admin。创建项目空间。",
)
def create_space(
    tenant_id: int,
    payload: SpaceCreateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    exists = session.exec(
        select(ProjectSpace).where(ProjectSpace.tenant_id == tenant_id, ProjectSpace.space_key == payload.space_key)
    ).first()
    if exists:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="space_key 已存在")
    item = ProjectSpace(tenant_id=tenant_id, space_key=payload.space_key, name=payload.name)
    session.add(item)
    session.commit()
    session.refresh(item)
    return SpaceOut.model_validate(item, from_attributes=True)


@router.put(
    "/tenants/{tenant_id}/spaces",
    response_model=SpaceOut,
    description="权限：tenant_admin（同项目）或 super_admin。更新项目空间（统一使用 PUT /spaces）。",
)
def update_space(
    tenant_id: int,
    payload: SpaceUpdateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    item = session.get(ProjectSpace, payload.id)
    if not item or item.tenant_id != tenant_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="空间不存在")
    item.name = payload.name
    item.updated_at = datetime.now(timezone.utc)
    session.add(item)
    session.commit()
    session.refresh(item)
    return SpaceOut.model_validate(item, from_attributes=True)


@router.post(
    "/tenants/{tenant_id}/spaces/delete",
    description="权限：tenant_admin（同项目）或 super_admin。批量删除项目空间。",
)
def delete_spaces(
    tenant_id: int,
    payload: SpaceDeleteIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    for sid in payload.ids:
        space = session.get(ProjectSpace, sid)
        if space and space.tenant_id == tenant_id:
            session.delete(space)
    session.commit()
    return {"deleted_ids": payload.ids, "behavior_key": "project-spaces/delete"}


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/questionnaires/questions",
    response_model=Page,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。问卷题目分页列表。",
)
def list_questions(
    tenant_id: int,
    space_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    stmt = select(QuestionnaireQuestion).where(
        QuestionnaireQuestion.tenant_id == tenant_id, QuestionnaireQuestion.project_space_id == space_id
    )
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    items = session.exec(stmt.order_by(QuestionnaireQuestion.sort_order).offset(skip).limit(limit)).all()
    return {"total": total, "items": [QuestionOut.model_validate(i, from_attributes=True).model_dump() for i in items]}


@router.post(
    "/tenants/{tenant_id}/spaces/{space_id}/questionnaires/questions",
    response_model=QuestionOut,
    description="权限：tenant_admin（同项目）或 super_admin。创建问卷题目。",
)
def create_question(
    tenant_id: int,
    space_id: int,
    payload: QuestionCreateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    space = session.get(ProjectSpace, space_id)
    if not space or space.tenant_id != tenant_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="空间不存在")
    item = QuestionnaireQuestion(tenant_id=tenant_id, project_space_id=space_id, **payload.model_dump())
    session.add(item)
    session.commit()
    session.refresh(item)
    return QuestionOut.model_validate(item, from_attributes=True)


@router.put(
    "/tenants/{tenant_id}/spaces/{space_id}/questionnaires/questions",
    response_model=QuestionOut,
    description="权限：tenant_admin（同项目）或 super_admin。更新问卷题目。",
)
def update_question(
    tenant_id: int,
    space_id: int,
    payload: QuestionUpdateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    item = session.get(QuestionnaireQuestion, payload.id)
    if not item or item.tenant_id != tenant_id or item.project_space_id != space_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="题目不存在")
    item.title = payload.title
    item.question_type = payload.question_type
    item.is_required = payload.is_required
    item.sort_order = payload.sort_order
    if payload.options_json is not None:
        item.options_json = payload.options_json
    item.updated_at = datetime.now(timezone.utc)
    session.add(item)
    session.commit()
    session.refresh(item)
    return QuestionOut.model_validate(item, from_attributes=True)


@router.post(
    "/tenants/{tenant_id}/spaces/{space_id}/questionnaires/questions/delete",
    description="权限：tenant_admin（同项目）或 super_admin。批量删除问卷题目。",
)
def delete_questions(
    tenant_id: int,
    space_id: int,
    payload: QuestionDeleteIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    for qid in payload.ids:
        item = session.get(QuestionnaireQuestion, qid)
        if item and item.tenant_id == tenant_id and item.project_space_id == space_id:
            session.delete(item)
    session.commit()
    return {"deleted_ids": payload.ids, "behavior_key": "questionnaire/questions/delete"}


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/taxonomy/nodes",
    response_model=Page,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。分类节点分页列表。",
)
def list_taxonomy_nodes(
    tenant_id: int,
    space_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    stmt = select(TaxonomyNode).where(TaxonomyNode.tenant_id == tenant_id, TaxonomyNode.project_space_id == space_id)
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    items = session.exec(stmt.order_by(TaxonomyNode.sort_order, TaxonomyNode.id).offset(skip).limit(limit)).all()
    return {"total": total, "items": [TaxonomyNodeOut.model_validate(i, from_attributes=True).model_dump() for i in items]}


@router.post(
    "/tenants/{tenant_id}/spaces/{space_id}/taxonomy/nodes",
    response_model=TaxonomyNodeOut,
    description="权限：tenant_admin（同项目）或 super_admin。创建分类节点。",
)
def create_taxonomy_node(
    tenant_id: int,
    space_id: int,
    payload: TaxonomyNodeCreateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    node = TaxonomyNode(tenant_id=tenant_id, project_space_id=space_id, **payload.model_dump())
    session.add(node)
    session.commit()
    session.refresh(node)
    return TaxonomyNodeOut.model_validate(node, from_attributes=True)


@router.put(
    "/tenants/{tenant_id}/spaces/{space_id}/taxonomy/nodes/{node_id}",
    response_model=TaxonomyNodeOut,
    description="权限：tenant_admin（同项目）或 super_admin。更新分类节点。",
)
def update_taxonomy_node(
    tenant_id: int,
    space_id: int,
    node_id: int,
    payload: TaxonomyNodeUpdateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    node = session.get(TaxonomyNode, node_id)
    if not node or node.tenant_id != tenant_id or node.project_space_id != space_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="分类节点不存在")
    for k, v in payload.model_dump().items():
        setattr(node, k, v)
    node.updated_at = datetime.now(timezone.utc)
    session.add(node)
    session.commit()
    session.refresh(node)
    return TaxonomyNodeOut.model_validate(node, from_attributes=True)


@router.delete(
    "/tenants/{tenant_id}/spaces/{space_id}/taxonomy/nodes/{node_id}",
    description="权限：tenant_admin（同项目）或 super_admin。删除分类节点。",
)
def delete_taxonomy_node(
    tenant_id: int,
    space_id: int,
    node_id: int,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    node = session.get(TaxonomyNode, node_id)
    if not node or node.tenant_id != tenant_id or node.project_space_id != space_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="分类节点不存在")
    session.delete(node)
    session.commit()
    return {"deleted_id": node_id, "behavior_key": "taxonomy-nodes/delete"}


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/relevance/rules",
    response_model=RelevanceRuleOut,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。读取相关性规则。",
)
def get_relevance_rule(
    tenant_id: int,
    space_id: int,
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    rule = session.exec(
        select(RelevanceRule).where(RelevanceRule.tenant_id == tenant_id, RelevanceRule.project_space_id == space_id)
    ).first()
    if not rule:
        return RelevanceRuleOut(tenant_id=tenant_id, project_space_id=space_id, expression="")
    return RelevanceRuleOut.model_validate(rule, from_attributes=True)


@router.put(
    "/tenants/{tenant_id}/spaces/{space_id}/relevance/rules",
    response_model=RelevanceRuleOut,
    description="权限：tenant_admin（同项目）或 super_admin。更新相关性规则。",
)
def put_relevance_rule(
    tenant_id: int,
    space_id: int,
    payload: RelevanceRuleIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    rule = session.exec(
        select(RelevanceRule).where(RelevanceRule.tenant_id == tenant_id, RelevanceRule.project_space_id == space_id)
    ).first()
    if not rule:
        rule = RelevanceRule(tenant_id=tenant_id, project_space_id=space_id, expression=payload.expression)
    else:
        rule.expression = payload.expression
        rule.updated_at = datetime.now(timezone.utc)
    session.add(rule)
    session.commit()
    session.refresh(rule)
    return RelevanceRuleOut.model_validate(rule, from_attributes=True)


@router.post(
    "/tenants/{tenant_id}/spaces/{space_id}/relevance/assessments",
    response_model=RelevanceAssessmentOut,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。提交相关性判定填报。",
)
def create_relevance_assessment(
    tenant_id: int,
    space_id: int,
    payload: RelevanceAssessmentCreateIn,
    current_user: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    submission = RelevanceAssessmentSubmission(
        tenant_id=tenant_id,
        project_space_id=space_id,
        submitter_user_id=current_user.id,
        conclusion=payload.conclusion,
    )
    session.add(submission)
    session.flush()
    for answer in payload.answers:
        session.add(
            RelevanceAssessmentAnswer(
                tenant_id=tenant_id,
                submission_id=submission.id,
                question_key=answer.question_key,
                answer_value=answer.answer_value,
            )
        )
    session.commit()
    session.refresh(submission)
    return RelevanceAssessmentOut(
        id=submission.id,
        tenant_id=submission.tenant_id,
        project_space_id=submission.project_space_id,
        submitter_user_id=submission.submitter_user_id,
        conclusion=submission.conclusion,
        created_at=submission.created_at,
        answers=payload.answers,
    )


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/relevance/assessments/{submission_id}",
    response_model=RelevanceAssessmentOut,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。查看单条相关性判定填报。",
)
def get_relevance_assessment(
    tenant_id: int,
    space_id: int,
    submission_id: int,
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    submission = session.get(RelevanceAssessmentSubmission, submission_id)
    if not submission or submission.tenant_id != tenant_id or submission.project_space_id != space_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="相关性填报不存在")
    answers = session.exec(
        select(RelevanceAssessmentAnswer).where(
            RelevanceAssessmentAnswer.tenant_id == tenant_id,
            RelevanceAssessmentAnswer.submission_id == submission_id,
        )
    ).all()
    return RelevanceAssessmentOut(
        id=submission.id,
        tenant_id=submission.tenant_id,
        project_space_id=submission.project_space_id,
        submitter_user_id=submission.submitter_user_id,
        conclusion=submission.conclusion,
        created_at=submission.created_at,
        answers=[RelevanceAssessmentAnswerIn(question_key=a.question_key, answer_value=a.answer_value) for a in answers],
    )


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/forms/lifecycle-field-config",
    response_model=Page,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。分页查询填报字段配置。",
)
def list_lifecycle_field_config(
    tenant_id: int,
    space_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    stmt = select(LifecycleFieldConfig).where(
        LifecycleFieldConfig.tenant_id == tenant_id, LifecycleFieldConfig.project_space_id == space_id
    )
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    items = session.exec(stmt.order_by(LifecycleFieldConfig.sort_order, LifecycleFieldConfig.id).offset(skip).limit(limit)).all()
    return {"total": total, "items": [LifecycleFieldConfigOut.model_validate(i, from_attributes=True).model_dump() for i in items]}


@router.post(
    "/tenants/{tenant_id}/spaces/{space_id}/forms/lifecycle-field-config",
    response_model=LifecycleFieldConfigOut,
    description="权限：tenant_admin（同项目）或 super_admin。新增填报字段配置。",
)
def create_lifecycle_field_config(
    tenant_id: int,
    space_id: int,
    payload: LifecycleFieldConfigCreateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    item = LifecycleFieldConfig(tenant_id=tenant_id, project_space_id=space_id, **payload.model_dump())
    session.add(item)
    session.commit()
    session.refresh(item)
    return LifecycleFieldConfigOut.model_validate(item, from_attributes=True)


@router.put(
    "/tenants/{tenant_id}/spaces/{space_id}/forms/lifecycle-field-config",
    response_model=LifecycleFieldConfigOut,
    description="权限：tenant_admin（同项目）或 super_admin。更新填报字段配置。",
)
def update_lifecycle_field_config(
    tenant_id: int,
    space_id: int,
    id: int,
    payload: LifecycleFieldConfigUpdateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    item = session.get(LifecycleFieldConfig, id)
    if not item or item.tenant_id != tenant_id or item.project_space_id != space_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="字段配置不存在")
    for key, value in payload.model_dump().items():
        setattr(item, key, value)
    item.updated_at = datetime.now(timezone.utc)
    session.add(item)
    session.commit()
    session.refresh(item)
    return LifecycleFieldConfigOut.model_validate(item, from_attributes=True)


@router.delete(
    "/tenants/{tenant_id}/spaces/{space_id}/forms/lifecycle-field-config",
    description="权限：tenant_admin（同项目）或 super_admin。删除填报字段配置。",
)
def delete_lifecycle_field_config(
    tenant_id: int,
    space_id: int,
    id: int,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    item = session.get(LifecycleFieldConfig, id)
    if not item or item.tenant_id != tenant_id or item.project_space_id != space_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="字段配置不存在")
    session.delete(item)
    session.commit()
    return {"deleted_id": id, "behavior_key": "lifecycle-field-config"}


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/sensitivity-levels",
    response_model=Page,
    description="权限：tenant_member。密级定义列表。",
)
def list_sensitivity_levels(
    tenant_id: int,
    space_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    stmt = select(SensitivityLevel).where(
        SensitivityLevel.tenant_id == tenant_id,
        SensitivityLevel.project_space_id == space_id,
    )
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    items = session.exec(stmt.order_by(SensitivityLevel.sort_order, SensitivityLevel.id).offset(skip).limit(limit)).all()
    return {
        "total": total,
        "items": [SensitivityLevelOut.model_validate(i, from_attributes=True).model_dump() for i in items],
    }


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/taxonomy-levels",
    response_model=Page,
    description="权限：tenant_member。分类树层级列表。",
)
def list_taxonomy_levels(
    tenant_id: int,
    space_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    stmt = select(TaxonomyLevel).where(
        TaxonomyLevel.tenant_id == tenant_id,
        TaxonomyLevel.project_space_id == space_id,
    )
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    items = session.exec(
        stmt.order_by(TaxonomyLevel.sort_order, TaxonomyLevel.level).offset(skip).limit(limit)
    ).all()
    return {
        "total": total,
        "items": [TaxonomyLevelOut.model_validate(i, from_attributes=True).model_dump() for i in items],
    }


@router.post(
    "/tenants/{tenant_id}/spaces/{space_id}/taxonomy-levels",
    response_model=TaxonomyLevelOut,
    description="权限：tenant_admin+。创建分类树层级。",
)
def create_taxonomy_level(
    tenant_id: int,
    space_id: int,
    payload: TaxonomyLevelCreateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    space = session.get(ProjectSpace, space_id)
    if not space or space.tenant_id != tenant_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="空间不存在")
    dup = session.exec(
        select(TaxonomyLevel).where(
            TaxonomyLevel.tenant_id == tenant_id,
            TaxonomyLevel.project_space_id == space_id,
            TaxonomyLevel.level == payload.level,
        )
    ).first()
    if dup:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=f"分类级 {payload.level} 已存在")
    row = TaxonomyLevel(
        tenant_id=tenant_id,
        project_space_id=space_id,
        level=payload.level,
        name=payload.name,
        description=payload.description,
        sort_order=payload.sort_order,
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    return TaxonomyLevelOut.model_validate(row, from_attributes=True)


@router.put(
    "/tenants/{tenant_id}/spaces/{space_id}/taxonomy-levels/{level_id}",
    response_model=TaxonomyLevelOut,
    description="权限：tenant_admin+。更新分类树层级。",
)
def update_taxonomy_level(
    tenant_id: int,
    space_id: int,
    level_id: int,
    payload: TaxonomyLevelUpdateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    row = session.get(TaxonomyLevel, level_id)
    if not row or row.tenant_id != tenant_id or row.project_space_id != space_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="分类树层级不存在")
    dup = session.exec(
        select(TaxonomyLevel).where(
            TaxonomyLevel.tenant_id == tenant_id,
            TaxonomyLevel.project_space_id == space_id,
            TaxonomyLevel.level == payload.level,
            TaxonomyLevel.id != level_id,
        )
    ).first()
    if dup:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=f"分类级 {payload.level} 已被占用")
    row.level = payload.level
    row.name = payload.name
    row.description = payload.description
    row.sort_order = payload.sort_order
    row.updated_at = datetime.now(timezone.utc)
    session.add(row)
    session.commit()
    session.refresh(row)
    return TaxonomyLevelOut.model_validate(row, from_attributes=True)


@router.delete(
    "/tenants/{tenant_id}/spaces/{space_id}/taxonomy-levels/{level_id}",
    description="权限：tenant_admin+。删除分类树层级。",
)
def delete_taxonomy_level(
    tenant_id: int,
    space_id: int,
    level_id: int,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    row = session.get(TaxonomyLevel, level_id)
    if not row or row.tenant_id != tenant_id or row.project_space_id != space_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="分类树层级不存在")
    session.delete(row)
    session.commit()
    return {"deleted_id": level_id, "behavior_key": "taxonomy-levels/delete"}


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/field-catalog",
    response_model=Page,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。字段主表分页列表。",
)
def list_field_catalog(
    tenant_id: int,
    space_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    stmt = select(FieldCatalogEntry).where(
        FieldCatalogEntry.tenant_id == tenant_id, FieldCatalogEntry.project_space_id == space_id
    )
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    items = session.exec(stmt.order_by(FieldCatalogEntry.id).offset(skip).limit(limit)).all()
    return {"total": total, "items": [FieldCatalogOut.model_validate(i, from_attributes=True).model_dump() for i in items]}


@router.post(
    "/tenants/{tenant_id}/spaces/{space_id}/field-catalog",
    response_model=FieldCatalogOut,
    description="权限：tenant_admin（同项目）或 super_admin。新增字段主表条目。",
)
def create_field_catalog(
    tenant_id: int,
    space_id: int,
    payload: FieldCatalogCreateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    item = FieldCatalogEntry(tenant_id=tenant_id, project_space_id=space_id, **payload.model_dump())
    session.add(item)
    session.commit()
    session.refresh(item)
    return FieldCatalogOut.model_validate(item, from_attributes=True)


@router.put(
    "/tenants/{tenant_id}/spaces/{space_id}/field-catalog/{entry_id}",
    response_model=FieldCatalogOut,
    description="权限：tenant_admin（同项目）或 super_admin。更新字段主表条目。",
)
def update_field_catalog(
    tenant_id: int,
    space_id: int,
    entry_id: int,
    payload: FieldCatalogUpdateIn,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    item = session.get(FieldCatalogEntry, entry_id)
    if not item or item.tenant_id != tenant_id or item.project_space_id != space_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="字段条目不存在")
    for k, v in payload.model_dump().items():
        setattr(item, k, v)
    item.updated_at = datetime.now(timezone.utc)
    session.add(item)
    session.commit()
    session.refresh(item)
    return FieldCatalogOut.model_validate(item, from_attributes=True)


@router.delete(
    "/tenants/{tenant_id}/spaces/{space_id}/field-catalog/{entry_id}",
    description="权限：tenant_admin（同项目）或 super_admin。删除字段主表条目。",
)
def delete_field_catalog(
    tenant_id: int,
    space_id: int,
    entry_id: int,
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    item = session.get(FieldCatalogEntry, entry_id)
    if not item or item.tenant_id != tenant_id or item.project_space_id != space_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="字段条目不存在")
    session.delete(item)
    session.commit()
    return {"deleted_id": entry_id, "behavior_key": "field-catalog"}


@router.post(
    "/tenants/{tenant_id}/spaces/{space_id}/field-catalog/batch-import",
    description="权限：tenant_admin（同项目）或 super_admin。批量导入字段主表，重复项跳过。",
)
def batch_import_field_catalog(
    tenant_id: int,
    space_id: int,
    items: list[FieldCatalogCreateIn],
    _: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    created, skipped = 0, 0
    for row in items:
        exists = session.exec(
            select(FieldCatalogEntry).where(
                FieldCatalogEntry.tenant_id == tenant_id,
                FieldCatalogEntry.project_space_id == space_id,
                FieldCatalogEntry.field_name == row.field_name,
            )
        ).first()
        if exists:
            skipped += 1
            continue
        session.add(FieldCatalogEntry(tenant_id=tenant_id, project_space_id=space_id, **row.model_dump()))
        created += 1
    session.commit()
    return {"created_count": created, "skipped_count": skipped, "behavior_key": "field-catalog"}


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/field-catalog/value-options",
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。返回字段主表可选值快照。",
)
def field_catalog_value_options(
    tenant_id: int,
    space_id: int,
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    items = session.exec(
        select(FieldCatalogEntry).where(
            FieldCatalogEntry.tenant_id == tenant_id, FieldCatalogEntry.project_space_id == space_id
        )
    ).all()
    return {
        "items": [{"id": i.id, "field_name": i.field_name, "identifier_key": i.identifier_key} for i in items],
        "behavior_key": "field-catalog",
    }


@router.post(
    "/tenants/{tenant_id}/spaces/{space_id}/field-requests",
    response_model=FieldRequestOut,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。创建字段新增申请。",
)
def create_field_request(
    tenant_id: int,
    space_id: int,
    payload: FieldRequestCreateIn,
    current_user: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    item = FieldRequest(
        tenant_id=tenant_id,
        project_space_id=space_id,
        requester_user_id=current_user.id,
        field_name=payload.field_name,
        reason=payload.reason,
    )
    session.add(item)
    session.commit()
    session.refresh(item)
    return FieldRequestOut.model_validate(item, from_attributes=True)


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/field-requests",
    response_model=Page,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。字段申请分页列表。",
)
def list_field_requests(
    tenant_id: int,
    space_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    status_filter: str | None = Query(default=None, alias="status"),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    stmt = select(FieldRequest).where(FieldRequest.tenant_id == tenant_id, FieldRequest.project_space_id == space_id)
    if status_filter:
        stmt = stmt.where(FieldRequest.status == status_filter)
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    items = session.exec(stmt.order_by(FieldRequest.id.desc()).offset(skip).limit(limit)).all()
    return {"total": total, "items": [FieldRequestOut.model_validate(i, from_attributes=True).model_dump() for i in items]}


@router.put(
    "/tenants/{tenant_id}/spaces/{space_id}/field-requests/{request_id}/review",
    response_model=FieldRequestOut,
    description="权限：tenant_admin（同项目）或 super_admin。审核字段新增申请。",
)
def review_field_request(
    tenant_id: int,
    space_id: int,
    request_id: int,
    payload: FieldRequestReviewIn,
    reviewer: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    if payload.status not in {"approved", "rejected"}:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="审核状态仅支持 approved/rejected")
    item = session.get(FieldRequest, request_id)
    if not item or item.tenant_id != tenant_id or item.project_space_id != space_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="字段申请不存在")
    item.status = payload.status
    item.review_comment = payload.review_comment
    item.reviewed_by_user_id = reviewer.id
    item.reviewed_at = datetime.now(timezone.utc)
    session.add(item)
    if payload.status == "approved":
        exists = session.exec(
            select(FieldCatalogEntry).where(
                FieldCatalogEntry.tenant_id == tenant_id,
                FieldCatalogEntry.project_space_id == space_id,
                FieldCatalogEntry.field_name == item.field_name,
            )
        ).first()
        if not exists:
            session.add(
                FieldCatalogEntry(
                    tenant_id=tenant_id,
                    project_space_id=space_id,
                    field_name=item.field_name,
                    identifier_key=item.field_name.lower().replace(" ", "_"),
                    data_type="string",
                )
            )
    session.commit()
    session.refresh(item)
    return FieldRequestOut.model_validate(item, from_attributes=True)


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/business-functions/options",
    response_model=Page,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。业务功能选项列表。",
)
def list_business_function_options(
    tenant_id: int,
    space_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    stmt = select(BusinessFunctionOption).where(
        BusinessFunctionOption.tenant_id == tenant_id, BusinessFunctionOption.project_space_id == space_id
    )
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    items = session.exec(stmt.order_by(BusinessFunctionOption.id.desc()).offset(skip).limit(limit)).all()
    return {"total": total, "items": [BusinessFunctionOptionOut.model_validate(i, from_attributes=True).model_dump() for i in items]}


@router.post(
    "/tenants/{tenant_id}/spaces/{space_id}/business-functions/option-requests",
    response_model=BusinessFunctionOptionRequestOut,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。提交业务功能选项申请。",
)
def create_business_option_request(
    tenant_id: int,
    space_id: int,
    payload: BusinessFunctionOptionRequestCreateIn,
    current_user: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    item = BusinessFunctionOptionRequest(
        tenant_id=tenant_id,
        project_space_id=space_id,
        requester_user_id=current_user.id,
        option_name=payload.option_name,
        reason=payload.reason,
    )
    session.add(item)
    session.commit()
    session.refresh(item)
    return BusinessFunctionOptionRequestOut.model_validate(item, from_attributes=True)


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/business-functions/option-requests",
    response_model=Page,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。业务功能选项申请列表。",
)
def list_business_option_requests(
    tenant_id: int,
    space_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    status_filter: str | None = Query(default=None, alias="status"),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    stmt = select(BusinessFunctionOptionRequest).where(
        BusinessFunctionOptionRequest.tenant_id == tenant_id,
        BusinessFunctionOptionRequest.project_space_id == space_id,
    )
    if status_filter:
        stmt = stmt.where(BusinessFunctionOptionRequest.status == status_filter)
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    items = session.exec(stmt.order_by(BusinessFunctionOptionRequest.id.desc()).offset(skip).limit(limit)).all()
    return {
        "total": total,
        "items": [BusinessFunctionOptionRequestOut.model_validate(i, from_attributes=True).model_dump() for i in items],
    }


@router.put(
    "/tenants/{tenant_id}/spaces/{space_id}/business-functions/option-requests/{request_id}/review",
    response_model=BusinessFunctionOptionRequestOut,
    description="权限：tenant_admin（同项目）或 super_admin。审核业务功能选项申请。",
)
def review_business_option_request(
    tenant_id: int,
    space_id: int,
    request_id: int,
    payload: BusinessFunctionOptionRequestReviewIn,
    reviewer: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    if payload.status not in {"approved", "rejected"}:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="审核状态仅支持 approved/rejected")
    item = session.get(BusinessFunctionOptionRequest, request_id)
    if not item or item.tenant_id != tenant_id or item.project_space_id != space_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="业务功能选项申请不存在")
    item.status = payload.status
    item.review_comment = payload.review_comment
    item.reviewed_by_user_id = reviewer.id
    session.add(item)
    if payload.status == "approved":
        existing = session.exec(
            select(BusinessFunctionOption).where(
                BusinessFunctionOption.tenant_id == tenant_id,
                BusinessFunctionOption.project_space_id == space_id,
                BusinessFunctionOption.option_name == item.option_name,
            )
        ).first()
        if not existing:
            session.add(
                BusinessFunctionOption(
                    tenant_id=tenant_id,
                    project_space_id=space_id,
                    option_name=item.option_name,
                    description=item.reason,
                    is_active=True,
                )
            )
    session.commit()
    session.refresh(item)
    return BusinessFunctionOptionRequestOut.model_validate(item, from_attributes=True)


@router.post(
    "/tenants/{tenant_id}/spaces/{space_id}/identifiers/suggest",
    response_model=IdentifierSuggestOut,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。根据种子文本建议 identifier_key。",
)
def suggest_identifier_key(
    tenant_id: int,
    space_id: int,
    payload: IdentifierSuggestIn,
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    space = session.get(ProjectSpace, space_id)
    if not space or space.tenant_id != tenant_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目空间不存在")
    key = suggest_identifier_key_from_seed(payload.seed)
    return IdentifierSuggestOut(suggested_key=key, behavior_key="suggest-identifier-key")


@router.post(
    "/tenants/{tenant_id}/spaces/{space_id}/field-usage-reports",
    response_model=FieldUsageReportOut,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。提交字段填报单。",
)
def create_field_usage_report(
    tenant_id: int,
    space_id: int,
    payload: FieldUsageReportCreateIn,
    current_user: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    space = session.get(ProjectSpace, space_id)
    if not space or space.tenant_id != tenant_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目空间不存在")
    if not payload.items:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="填报项不能为空")
    report = FieldUsageReport(
        tenant_id=tenant_id,
        project_space_id=space_id,
        submitter_user_id=current_user.id,
        title=payload.title,
        status="pending_review",
    )
    session.add(report)
    session.flush()
    for row in payload.items:
        session.add(
            FieldUsageReportItem(
                tenant_id=tenant_id,
                report_id=report.id,
                field_name=row.field_name,
                value_text=row.value_text or "",
            )
        )
    session.commit()
    session.refresh(report)
    items = session.exec(
        select(FieldUsageReportItem).where(
            FieldUsageReportItem.tenant_id == tenant_id,
            FieldUsageReportItem.report_id == report.id,
        )
    ).all()
    return FieldUsageReportOut(
        id=report.id,
        tenant_id=report.tenant_id,
        project_space_id=report.project_space_id,
        submitter_user_id=report.submitter_user_id,
        title=report.title,
        status=report.status,
        review_comment=report.review_comment,
        created_at=report.created_at,
        items=[FieldUsageReportItemOut.model_validate(i, from_attributes=True) for i in items],
        behavior_key="field-usage-reports",
    )


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/field-usage-reports",
    response_model=Page,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。字段填报单分页列表。",
)
def list_field_usage_reports(
    tenant_id: int,
    space_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    status_filter: str | None = Query(default=None, alias="status"),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    space = session.get(ProjectSpace, space_id)
    if not space or space.tenant_id != tenant_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目空间不存在")
    stmt = select(FieldUsageReport).where(
        FieldUsageReport.tenant_id == tenant_id,
        FieldUsageReport.project_space_id == space_id,
    )
    if status_filter:
        stmt = stmt.where(FieldUsageReport.status == status_filter)
    total = session.exec(select(func.count()).select_from(stmt.subquery())).one()
    reports = session.exec(stmt.order_by(FieldUsageReport.id.desc()).offset(skip).limit(limit)).all()
    ids = [r.id for r in reports]
    counts: dict[int, int] = {}
    if ids:
        for it in session.exec(select(FieldUsageReportItem).where(FieldUsageReportItem.report_id.in_(ids))).all():
            counts[it.report_id] = counts.get(it.report_id, 0) + 1
    items_out = [
        FieldUsageReportListItemOut(
            id=r.id,
            tenant_id=r.tenant_id,
            project_space_id=r.project_space_id,
            submitter_user_id=r.submitter_user_id,
            title=r.title,
            status=r.status,
            created_at=r.created_at,
            item_count=counts.get(r.id, 0),
        ).model_dump()
        for r in reports
    ]
    return {"total": total, "items": items_out}


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/field-usage-reports/export",
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。导出字段填报明细为 CSV。",
)
def export_field_usage_reports(
    tenant_id: int,
    space_id: int,
    status_filter: str | None = Query(default=None, alias="status"),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    space = session.get(ProjectSpace, space_id)
    if not space or space.tenant_id != tenant_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目空间不存在")
    r_stmt = select(FieldUsageReport).where(
        FieldUsageReport.tenant_id == tenant_id,
        FieldUsageReport.project_space_id == space_id,
    )
    if status_filter:
        r_stmt = r_stmt.where(FieldUsageReport.status == status_filter)
    reports = session.exec(r_stmt.order_by(FieldUsageReport.id.desc())).all()
    report_ids = [r.id for r in reports]
    items_by_report: dict[int, list[FieldUsageReportItem]] = {rid: [] for rid in report_ids}
    if report_ids:
        for it in session.exec(select(FieldUsageReportItem).where(FieldUsageReportItem.report_id.in_(report_ids))).all():
            items_by_report.setdefault(it.report_id, []).append(it)
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow(["report_id", "title", "report_status", "field_name", "value_text", "created_at"])
    for r in reports:
        its = items_by_report.get(r.id, [])
        if not its:
            w.writerow([r.id, r.title or "", r.status, "", "", r.created_at.isoformat()])
        else:
            for it in its:
                w.writerow([r.id, r.title or "", r.status, it.field_name, it.value_text, r.created_at.isoformat()])
    data = "\ufeff" + buf.getvalue()
    return StreamingResponse(
        iter([data.encode("utf-8")]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="field-usage-reports.csv"'},
    )


@router.post(
    "/tenants/{tenant_id}/spaces/{space_id}/field-usage-reports/{report_id}/review",
    response_model=FieldUsageReportOut,
    description="权限：tenant_admin（同项目）或 super_admin。审核字段填报单。",
)
def review_field_usage_report(
    tenant_id: int,
    space_id: int,
    report_id: int,
    payload: FieldUsageReportReviewIn,
    reviewer: User = Depends(require_tenant_admin),
    session: Session = Depends(get_session),
):
    if payload.status not in {"approved", "rejected"}:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="审核状态仅支持 approved/rejected")
    report = session.get(FieldUsageReport, report_id)
    if not report or report.tenant_id != tenant_id or report.project_space_id != space_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="填报单不存在")
    report.status = payload.status
    report.review_comment = payload.review_comment
    report.reviewed_by_user_id = reviewer.id
    report.reviewed_at = datetime.now(timezone.utc)
    session.add(report)
    session.commit()
    session.refresh(report)
    items = session.exec(
        select(FieldUsageReportItem).where(
            FieldUsageReportItem.tenant_id == tenant_id,
            FieldUsageReportItem.report_id == report.id,
        )
    ).all()
    return FieldUsageReportOut(
        id=report.id,
        tenant_id=report.tenant_id,
        project_space_id=report.project_space_id,
        submitter_user_id=report.submitter_user_id,
        title=report.title,
        status=report.status,
        review_comment=report.review_comment,
        created_at=report.created_at,
        items=[FieldUsageReportItemOut.model_validate(i, from_attributes=True) for i in items],
        behavior_key="field-usage-reports",
    )


@router.get(
    "/tenants/{tenant_id}/spaces/{space_id}/work-orders",
    response_model=Page,
    description="权限：tenant_member / tenant_admin（同项目）或 super_admin。合并待办：待审核字段申请、业务选项申请、字段填报单。",
)
def list_work_orders(
    tenant_id: int,
    space_id: int,
    skip: int = 0,
    limit: int = Query(default=20),
    _: User = Depends(require_tenant_member),
    session: Session = Depends(get_session),
):
    limit = page_limit(limit)
    space = session.get(ProjectSpace, space_id)
    if not space or space.tenant_id != tenant_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目空间不存在")
    frs = session.exec(
        select(FieldRequest).where(
            FieldRequest.tenant_id == tenant_id,
            FieldRequest.project_space_id == space_id,
            FieldRequest.status == "pending",
        )
    ).all()
    ors = session.exec(
        select(BusinessFunctionOptionRequest).where(
            BusinessFunctionOptionRequest.tenant_id == tenant_id,
            BusinessFunctionOptionRequest.project_space_id == space_id,
            BusinessFunctionOptionRequest.status == "pending",
        )
    ).all()
    urs = session.exec(
        select(FieldUsageReport).where(
            FieldUsageReport.tenant_id == tenant_id,
            FieldUsageReport.project_space_id == space_id,
            FieldUsageReport.status == "pending_review",
        )
    ).all()
    merged: list[tuple[datetime, str, object]] = []
    for x in frs:
        merged.append((x.created_at, "field_request", x))
    for x in ors:
        merged.append((x.created_at, "business_option_request", x))
    for x in urs:
        merged.append((x.created_at, "field_usage_report", x))
    merged.sort(key=lambda t: t[0], reverse=True)
    total = len(merged)
    page = merged[skip : skip + limit]
    items_out: list[dict] = []
    for _created_at, kind, obj in page:
        if kind == "field_request":
            items_out.append(
                {
                    "order_kind": "field_request",
                    "id": obj.id,
                    "status": obj.status,
                    "created_at": obj.created_at,
                    "summary": obj.field_name,
                    "requester_user_id": obj.requester_user_id,
                }
            )
        elif kind == "business_option_request":
            items_out.append(
                {
                    "order_kind": "business_option_request",
                    "id": obj.id,
                    "status": obj.status,
                    "created_at": obj.created_at,
                    "summary": obj.option_name,
                    "requester_user_id": obj.requester_user_id,
                }
            )
        else:
            items_out.append(
                {
                    "order_kind": "field_usage_report",
                    "id": obj.id,
                    "status": obj.status,
                    "created_at": obj.created_at,
                    "summary": obj.title or f"填报单#{obj.id}",
                    "submitter_user_id": obj.submitter_user_id,
                }
            )
    return {"total": total, "items": items_out}
