"""文档资源：文件存储 + Excel 导入/导出（首版实现 field_catalog 等核心模块）。"""

from __future__ import annotations

import hashlib
import json
import re
import uuid
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

from app.core.config import settings
from app.models import FieldCatalogEntry, FieldClassGrade, LifecycleFieldConfig, TaxonomyNode
from app.models_portal import (
    BusinessFunction,
    DocumentResource,
    DocumentTransferJob,
    SensitivityLevel,
    SubmissionTask,
    TaxonomyLevel,
)
from app.services.document_registry import get_module_spec


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _tenant_upload_dir(tenant_id: int) -> Path:
    root = Path(settings.upload_root)
    path = root / str(tenant_id)
    path.mkdir(parents=True, exist_ok=True)
    return path


def save_upload_file(tenant_id: int, original_name: str, content: bytes) -> tuple[str, str]:
    digest = hashlib.sha256(content).hexdigest()
    ext = Path(original_name).suffix or ".bin"
    rel = f"{uuid.uuid4().hex}{ext}"
    full = _tenant_upload_dir(tenant_id) / rel
    full.write_bytes(content)
    return rel, digest


def resolve_storage_path(tenant_id: int, storage_path: str) -> Path:
    return _tenant_upload_dir(tenant_id) / storage_path


def create_document_resource(
    session: Session,
    *,
    tenant_id: int,
    project_space_id: int | None,
    resource_kind: str,
    module_key: str | None,
    title: str,
    description: str | None,
    original_file_name: str,
    mime_type: str,
    content: bytes,
    uploaded_by_user_id: int,
    tags_json: str | None = None,
    regulation_meta_json: str | None = None,
) -> DocumentResource:
    if len(content) > settings.upload_max_bytes:
        raise ValueError("文件超过大小限制")
    storage_path, checksum = save_upload_file(tenant_id, original_file_name, content)
    row = DocumentResource(
        tenant_id=tenant_id,
        project_space_id=project_space_id,
        resource_kind=resource_kind,
        module_key=module_key,
        title=title,
        description=description,
        original_file_name=original_file_name,
        mime_type=mime_type,
        file_size_bytes=len(content),
        storage_path=storage_path,
        checksum_sha256=checksum,
        tags_json=tags_json,
        regulation_meta_json=regulation_meta_json,
        uploaded_by_user_id=uploaded_by_user_id,
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    return row


def _slug_identifier(label: str) -> str:
    base = re.sub(r"[^a-zA-Z0-9_]+", "_", label.strip().lower())
    base = base.strip("_") or "field"
    return base[:64]


def build_module_template_xlsx(module_key: str) -> bytes:
    spec = get_module_spec(module_key)
    if not spec:
        raise ValueError("未知模块")
    wb = Workbook()
    ws = wb.active
    ws.title = module_key[:31]

    if module_key == "field_catalog":
        ws.append(["field_name", "description", "identifier_key", "data_type", "taxonomy_code", "source_system"])
        ws.append(["用户手机号", "示例", "user_mobile", "string", "", "CRM"])
    elif module_key == "sensitivity_level":
        ws.append(["code", "label", "description", "sort_order"])
        ws.append(["internal", "内部", "", "1"])
    elif module_key == "taxonomy_level":
        ws.append(["level", "name", "description", "sort_order"])
        ws.append(["0", "根层级", "", "0"])
    elif module_key == "taxonomy_node":
        ws.append(["code", "name", "parent_code", "sort_order"])
        ws.append(["L1", "一级分类", "", "0"])
    elif module_key == "field_class_grade":
        ws.append(["field_name", "grade_label", "notes"])
        ws.append(["用户手机号", "内部", ""])
    elif module_key == "field_taxonomy_binding":
        ws.append(["field_name", "taxonomy_code"])
        ws.append(["用户手机号", "L1"])
    elif module_key == "lifecycle_field_config":
        ws.append(["field_key", "field_label", "field_type", "required", "sort_order", "help_text"])
        ws.append(["custom_note", "备注", "textarea", "0", "10", ""])
    elif module_key == "business_function":
        ws.append(["function_key", "name", "description", "requires_fo_binding", "sort_order"])
        ws.append(["field_usage", "字段填报", "", "1", "1"])
    elif module_key == "relevance_questionnaire":
        ws.append(["key", "title", "sort_order", "options_csv"])
        ws.append(["q1", "是否处理个人信息？", "0", "是;否"])
    else:
        ws.append(["note"])
        ws.append(["该模块模板将在后续版本完善"])

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def export_module_xlsx(session: Session, tenant_id: int, space_id: int, module_key: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = module_key[:31]

    if module_key == "field_catalog":
        ws.append(["field_name", "description", "identifier_key", "data_type", "taxonomy_code", "source_system"])
        rows = session.exec(
            select(FieldCatalogEntry).where(
                FieldCatalogEntry.tenant_id == tenant_id,
                FieldCatalogEntry.project_space_id == space_id,
            )
        ).all()
        for r in rows:
            ws.append([r.field_name, r.description or "", r.identifier_key, r.data_type, r.taxonomy_code or "", r.source_system or ""])
    elif module_key == "sensitivity_level":
        ws.append(["code", "label", "description", "sort_order"])
        for r in session.exec(
            select(SensitivityLevel).where(
                SensitivityLevel.tenant_id == tenant_id,
                SensitivityLevel.project_space_id == space_id,
            )
        ).all():
            ws.append([r.code, r.label, r.description or "", r.sort_order])
    elif module_key == "taxonomy_level":
        ws.append(["level", "name", "description", "sort_order"])
        for r in session.exec(
            select(TaxonomyLevel).where(
                TaxonomyLevel.tenant_id == tenant_id,
                TaxonomyLevel.project_space_id == space_id,
            )
        ).all():
            ws.append([r.level, r.name, r.description or "", r.sort_order])
    elif module_key == "taxonomy_node":
        ws.append(["code", "name", "parent_code", "sort_order"])
        nodes = session.exec(
            select(TaxonomyNode).where(
                TaxonomyNode.tenant_id == tenant_id,
                TaxonomyNode.project_space_id == space_id,
            )
        ).all()
        id_to_code = {n.id: n.code for n in nodes}
        for n in nodes:
            parent_code = id_to_code.get(n.parent_id, "") if n.parent_id else ""
            ws.append([n.code, n.name, parent_code, n.sort_order])
    elif module_key == "field_class_grade":
        ws.append(["field_name", "grade_label", "notes"])
        grades = session.exec(
            select(FieldClassGrade).where(
                FieldClassGrade.tenant_id == tenant_id,
                FieldClassGrade.project_space_id == space_id,
            )
        ).all()
        for g in grades:
            entry = session.get(FieldCatalogEntry, g.field_catalog_entry_id)
            ws.append([entry.field_name if entry else str(g.field_catalog_entry_id), g.grade_label, g.notes or ""])
    elif module_key == "field_taxonomy_binding":
        ws.append(["field_name", "taxonomy_code"])
        for e in session.exec(
            select(FieldCatalogEntry).where(
                FieldCatalogEntry.tenant_id == tenant_id,
                FieldCatalogEntry.project_space_id == space_id,
            )
        ).all():
            ws.append([e.field_name, e.taxonomy_code or ""])
    elif module_key == "lifecycle_field_config":
        ws.append(["field_key", "field_label", "field_type", "required", "sort_order", "help_text"])
        for c in session.exec(
            select(LifecycleFieldConfig).where(
                LifecycleFieldConfig.tenant_id == tenant_id,
                LifecycleFieldConfig.project_space_id == space_id,
            )
        ).all():
            ws.append([c.field_key, c.field_label, c.field_type, "1" if c.is_required else "0", c.sort_order, ""])
    elif module_key == "business_function":
        ws.append(["function_key", "name", "description", "requires_fo_binding", "sort_order"])
        for f in session.exec(
            select(BusinessFunction).where(
                BusinessFunction.tenant_id == tenant_id,
                BusinessFunction.project_space_id == space_id,
            )
        ).all():
            ws.append([f.function_key, f.name, f.description or "", "1" if f.requires_fo_binding else "0", f.sort_order])
    elif module_key == "submission_task":
        ws.append(["id", "title", "function_key", "status", "fo_fill_status", "audit_status"])
        tasks = session.exec(
            select(SubmissionTask).where(
                SubmissionTask.tenant_id == tenant_id,
                SubmissionTask.project_space_id == space_id,
            )
        ).all()
        fn_map = {
            f.id: f.function_key
            for f in session.exec(
                select(BusinessFunction).where(
                    BusinessFunction.tenant_id == tenant_id,
                    BusinessFunction.project_space_id == space_id,
                )
            ).all()
        }
        for t in tasks:
            ws.append([t.id, t.title, fn_map.get(t.business_function_id, ""), t.status, t.fo_fill_status, t.audit_status or ""])
    else:
        ws.append(["message"])
        ws.append(["该模块导出尚未实现"])

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def import_field_catalog_from_xlsx(
    session: Session, tenant_id: int, space_id: int, content: bytes
) -> dict:
    wb = load_workbook(filename=BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"total_rows": 0, "created": 0, "updated": 0, "skipped": 0, "errors": [{"row": 0, "message": "空文件"}]}

    header = [str(c or "").strip() for c in rows[0]]
    header_map = {name: idx for idx, name in enumerate(header)}
    if "field_name" not in header_map:
        return {"total_rows": 0, "created": 0, "updated": 0, "skipped": 0, "errors": [{"row": 1, "message": "缺少 field_name 列"}]}

    created = updated = skipped = 0
    errors: list[dict] = []
    total = max(len(rows) - 1, 0)

    for ridx, row in enumerate(rows[1:], start=2):
        def cell(name: str) -> str:
            idx = header_map.get(name)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        field_name = cell("field_name")
        if not field_name:
            skipped += 1
            errors.append({"row": ridx, "message": "field_name 为空"})
            continue

        existing = session.exec(
            select(FieldCatalogEntry).where(
                FieldCatalogEntry.tenant_id == tenant_id,
                FieldCatalogEntry.project_space_id == space_id,
                FieldCatalogEntry.field_name == field_name,
            )
        ).first()

        identifier_key = cell("identifier_key") or _slug_identifier(field_name)
        description = cell("description") or None
        data_type = cell("data_type") or "string"
        taxonomy_code = cell("taxonomy_code") or None
        source_system = cell("source_system") or None

        if existing:
            existing.description = description
            existing.identifier_key = identifier_key or existing.identifier_key
            existing.data_type = data_type
            existing.taxonomy_code = taxonomy_code
            existing.source_system = source_system
            session.add(existing)
            updated += 1
        else:
            session.add(
                FieldCatalogEntry(
                    tenant_id=tenant_id,
                    project_space_id=space_id,
                    field_name=field_name,
                    description=description,
                    identifier_key=identifier_key,
                    data_type=data_type,
                    taxonomy_code=taxonomy_code,
                    source_system=source_system,
                )
            )
            created += 1

    session.commit()
    return {"total_rows": total, "created": created, "updated": updated, "skipped": skipped, "errors": errors}


def _upsert_taxonomy_level(session: Session, tenant_id: int, space_id: int, row: dict) -> None:
    level = int(row.get("level") or 0)
    name = str(row.get("name") or "").strip()
    if not name:
        raise ValueError("name 不能为空")
    sort_order = int(row.get("sort_order") or level)
    description = row.get("description") or None
    existing = session.exec(
        select(TaxonomyLevel).where(
            TaxonomyLevel.tenant_id == tenant_id,
            TaxonomyLevel.project_space_id == space_id,
            TaxonomyLevel.level == level,
        )
    ).first()
    if existing:
        existing.name = name
        existing.description = description
        existing.sort_order = sort_order
        session.add(existing)
    else:
        session.add(
            TaxonomyLevel(
                tenant_id=tenant_id,
                project_space_id=space_id,
                level=level,
                name=name,
                description=description,
                sort_order=sort_order,
            )
        )


def _upsert_taxonomy_node(session: Session, tenant_id: int, space_id: int, row: dict) -> None:
    code = str(row.get("code") or "").strip()
    name = str(row.get("name") or "").strip()
    if not code or not name:
        raise ValueError("code/name 不能为空")
    parent_code = str(row.get("parent_code") or "").strip()
    sort_order = int(row.get("sort_order") or 0)
    parent_id = None
    if parent_code:
        parent = session.exec(
            select(TaxonomyNode).where(
                TaxonomyNode.tenant_id == tenant_id,
                TaxonomyNode.project_space_id == space_id,
                TaxonomyNode.code == parent_code,
            )
        ).first()
        if not parent:
            raise ValueError(f"未找到上级节点 code: {parent_code}")
        parent_id = parent.id
    existing = session.exec(
        select(TaxonomyNode).where(
            TaxonomyNode.tenant_id == tenant_id,
            TaxonomyNode.project_space_id == space_id,
            TaxonomyNode.code == code,
        )
    ).first()
    if existing:
        existing.name = name
        existing.parent_id = parent_id
        existing.sort_order = sort_order
        session.add(existing)
    else:
        session.add(
            TaxonomyNode(
                tenant_id=tenant_id,
                project_space_id=space_id,
                code=code,
                name=name,
                parent_id=parent_id,
                sort_order=sort_order,
            )
        )


def import_module_from_xlsx(session: Session, tenant_id: int, space_id: int, module_key: str, content: bytes) -> dict:
    if module_key == "field_catalog":
        return import_field_catalog_from_xlsx(session, tenant_id, space_id, content)
    if module_key == "sensitivity_level":
        return _import_simple_rows(
            session,
            content,
            required=["code", "label"],
            handler=lambda row: _upsert_sensitivity(session, tenant_id, space_id, row),
        )
    if module_key == "taxonomy_level":
        return _import_simple_rows(
            session,
            content,
            required=["level", "name"],
            handler=lambda row: _upsert_taxonomy_level(session, tenant_id, space_id, row),
        )
    if module_key == "taxonomy_node":
        return _import_simple_rows(
            session,
            content,
            required=["code", "name"],
            handler=lambda row: _upsert_taxonomy_node(session, tenant_id, space_id, row),
        )
    raise ValueError(f"模块 {module_key} 的导入尚未实现")


def _upsert_sensitivity(session: Session, tenant_id: int, space_id: int, row: dict) -> None:
    code = row.get("code", "").strip()
    label = row.get("label", "").strip()
    if not code or not label:
        raise ValueError("code/label 不能为空")
    sort_order = int(row.get("sort_order") or 0)
    description = row.get("description") or None
    existing = session.exec(
        select(SensitivityLevel).where(
            SensitivityLevel.tenant_id == tenant_id,
            SensitivityLevel.project_space_id == space_id,
            SensitivityLevel.code == code,
        )
    ).first()
    if existing:
        existing.label = label
        existing.description = description
        existing.sort_order = sort_order
        session.add(existing)
    else:
        session.add(
            SensitivityLevel(
                tenant_id=tenant_id,
                project_space_id=space_id,
                code=code,
                label=label,
                description=description,
                sort_order=sort_order,
            )
        )


def _import_simple_rows(session: Session, content: bytes, required: list[str], handler) -> dict:
    wb = load_workbook(filename=BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"total_rows": 0, "created": 0, "updated": 0, "skipped": 0, "errors": []}
    header = [str(c or "").strip() for c in rows[0]]
    header_map = {name: idx for idx, name in enumerate(header)}
    for col in required:
        if col not in header_map:
            return {"total_rows": 0, "created": 0, "updated": 0, "skipped": 0, "errors": [{"row": 1, "message": f"缺少 {col} 列"}]}

    created = updated = skipped = 0
    errors: list[dict] = []
    total = max(len(rows) - 1, 0)

    for ridx, row in enumerate(rows[1:], start=2):
        data = {}
        for name, idx in header_map.items():
            data[name] = str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ""
        try:
            handler(data)
            created += 1
        except Exception as exc:  # noqa: BLE001 — 行级错误汇总
            skipped += 1
            errors.append({"row": ridx, "message": str(exc)})

    session.commit()
    return {"total_rows": total, "created": created, "updated": updated, "skipped": skipped, "errors": errors}


def run_export_job(session: Session, job: DocumentTransferJob) -> DocumentTransferJob:
    job.status = "running"
    job.started_at = utc_now()
    session.add(job)
    session.commit()
    try:
        content = export_module_xlsx(session, job.tenant_id, job.project_space_id, job.module_key)
        spec = get_module_spec(job.module_key)
        filename = spec.template_filename if spec else f"{job.module_key}_export.xlsx"
        resource = create_document_resource(
            session,
            tenant_id=job.tenant_id,
            project_space_id=job.project_space_id,
            resource_kind="export_artifact",
            module_key=job.module_key,
            title=f"{job.module_key} 导出",
            description=None,
            original_file_name=filename,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=content,
            uploaded_by_user_id=job.created_by_user_id,
        )
        job.result_resource_id = resource.id
        job.result_summary_json = json.dumps({"rows_exported": "see file"}, ensure_ascii=False)
        job.status = "succeeded"
        job.completed_at = utc_now()
    except Exception as exc:  # noqa: BLE001
        job.status = "failed"
        job.error_message = str(exc)
        job.completed_at = utc_now()
    session.add(job)
    session.commit()
    session.refresh(job)
    return job


def run_import_job(session: Session, job: DocumentTransferJob, file_content: bytes) -> DocumentTransferJob:
    job.status = "running"
    job.started_at = utc_now()
    session.add(job)
    session.commit()
    try:
        summary = import_module_from_xlsx(session, job.tenant_id, job.project_space_id, job.module_key, file_content)
        job.result_summary_json = json.dumps(summary, ensure_ascii=False)
        job.status = "succeeded" if not summary.get("errors") else "succeeded"
        job.completed_at = utc_now()
    except Exception as exc:  # noqa: BLE001
        job.status = "failed"
        job.error_message = str(exc)
        job.completed_at = utc_now()
    session.add(job)
    session.commit()
    session.refresh(job)
    return job
